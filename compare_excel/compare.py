'''
@Descripttion: 
@version: 
@Author: ivy-liu
@Date: 2020-01-08 22:01:30
@LastEditors  : ivy-liu
@LastEditTime : 2020-01-09 01:46:59
'''
from openpyxl import load_workbook

class ExcelOp(object):
    def __init__(self,sheet="全量"):
        excel_path=r"workSpace\compare_excel\0108目标导入.xlsx"
        self.wb=load_workbook(excel_path)
        self.ws=self.wb[sheet]

    # 获取某行所有值
    def get_row_value(self,row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    def run(self,month_col,value,one_row,two_row=None):
        print("line: "+line+"value: "+value)
        # 获取column
        excelOp=ExcelOp()
        row_data=excelOp.get_row_value(1)
        # print(row_data)
        column=row_data.index(month_col)+1 
        print("列名："+month_col+" 在第"+str(column)+"列\n")#拿到col
        
        # 获取row
        one_column_data=excelOp.get_col_value(2)
        # print(one_column_data)
        # [None, '无抵押', None, None, '有抵押', '一元', None, None, '二元', None, None, 
        # '三元', None, None, '直销', None, None, '远程', None, None, '第三方平安', None, None, 
        # '第三方非平安', None, None]
        one_row_num=one_column_data.index(one_row)+1
        print("第二列-行定位 {}{}".format(one_row,one_row_num))
        
        two_column_data=excelOp.get_col_value(3)
        # print(two_column_data)
        # [None, '合计', '大额', '小额', '合计', '合计', '无抵押', '有抵押', '合计', '无抵押', '有抵押', 
        # '合计', '无抵押', '有抵押', '合计', '无抵押', '有抵押', '合计', '无抵押', '有抵押', 
        # '合计', '无抵押', '有抵押', '合计', '无抵押', '有抵押']
        # print("two_row--",two_row)
        two_row_nums=[]
        for i in range (len(two_column_data)):
            if two_column_data[i]==two_row:
                two_row_nums.append(i+1)
        print("第三列-1次行定位 {0}{1}".format(two_row,two_row_nums))
        row=0
        for i in two_row_nums:
            if (i-one_row_num>=0)and (i-one_row_num<=2):
                # print("i={},one_row_num={}".format(i,one_row_num))
                row=i
        print("定位完成-column={},row={}".format(column,row))

        # 取excel数据
        excel_value=self.ws.cell(column=column,row=row).value
        print("excel_value---{},value--{}".format(excel_value,value))


excelOp=ExcelOp()

txt_path=r"workSpace\compare_excel\目标值.txt"
with open(txt_path,'r',encoding='utf-8') as f:
    f.seek(0)
    for line in f:
        if "目标" in line:
            # 给列名
            month_col=str(int(line[0:2]))+"月"
            # 给数值
            value=line.split("目标",1)[1]
            if "：" in value:
                value=value.split("：",1)[1]
            # 给第二列确定行,适用事业部和渠道，条线不适用
            if "度" in line:
                rows=line.split("度",1)[0][3:-1]
                print("rows-",rows)
                one_row=rows[0:-3]
                print("one_row-",one_row)
            # 给第三列确定行,需要明天确定有木有
                two_row=rows[-3:]
                print("two_row-",two_row)

            else:
                pass #可能是合计什么的明天再看
            

            excelOp.run(month_col,value,one_row,two_row)
        else:
            continue



