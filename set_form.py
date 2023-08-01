from openpyxl.styles import Font,colors
from openpyxl.utils import get_column_letter

def set_title(sheet,title,color):
    title_font=Font(name='幼圆',size=16,bold=True,italic=True,color=color)
    len_title=len(title)
    for i in range (0,len_title):
        # print('title[i-1]---',title[i-1])
        sheet.cell(row=1,column=i+1).value=title[i]
        sheet.cell(row=1,column=i+1).font=title_font
    sheet.sheet_properties.tabColor=color #设置工作表标签颜色
    
def set_wid(sheet):
     #获取每列宽度 
     col_widths=[]
     #每列 
     for col in sheet.columns:
         #print(col)
         #每行 
         col_width=10
         for j in range(len(col)):
             #print(len(str(col[j].value)))
             if len(str(col[j].value))>=100:
                 col_width=100
             elif len(str(col[j].value))>col_width:
                 col_width=len(str(col[j].value))
             else:
                 pass
         col_widths.append(col_width)
      
      #设置列宽
      for(i,j) in zip(col_widths,range(0,sheet.max_column)):
          print(i,j)
          # 根据列的数字返回字母 
          # print(get_column_letter(2)) #B 
          col_letter=get_column_letter(j+1)
          sheet.column_dimensions[col_letter].width=i